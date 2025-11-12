#!/usr/bin/env python3
"""
Create AI Agent Portfolio Excel Spreadsheet
Converts the Google Apps Script to a Python script that generates an Excel file
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def create_agent_portfolio_excel(output_dir='./sheets'):
    """Create the Agent Portfolio Excel workbook"""

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Colors (hex codes from the original script)
    colors = {
        'purpleGrad': '8B5CF6',
        'lightPurple': 'F3E8FF',
        'lightBlue': 'DBEAFE',
        'darkGray': '374151',
        'white': 'FFFFFF',
        'purple': 'A78BFA',
        'blue': '60A5FA',
        'green': '34D399',
        'orange': 'FB923C',
        'cyan': '22D3EE',
        'pink': 'F472B6',
        'yellow': 'FBBF24',
        'red': 'EF4444',
        'indigo': '6366F1',
        'lightGray': 'E5E7EB',
        'black': '000000',
        'amber': 'F59E0B',
        'gray700': '6B7280',
        'blue500': '3B82F6',
        'green500': '10B981',
        'red600': 'DC2626'
    }

    # Create sheets
    main = wb.create_sheet('Agent Portfolio', 0)
    quick = wb.create_sheet('Quick Wins', 1)
    roadmap = wb.create_sheet('Build Roadmap', 2)

    # === MAIN SHEET HEADER SECTION ===
    # Row 1: Title
    main.merge_cells('A1:J1')
    main['A1'] = "🤖 MADHAVAN'S AI AGENT FORCE\nIntelligent Agents to 10x Your\nExecutive Leverage"
    main['A1'].font = Font(size=24, bold=True, color=colors['white'])
    main['A1'].fill = PatternFill(start_color=colors['purpleGrad'], end_color=colors['purpleGrad'], fill_type='solid')
    main['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    main.row_dimensions[1].height = 72

    # Row 2: Summary
    main.merge_cells('A2:J2')
    main['A2'] = "Total Time Saved: 67 hours/week\nYour Current Week: 80 hours →\nFuture Week: 40 hours strategic"
    main['A2'].font = Font(size=14)
    main['A2'].fill = PatternFill(start_color=colors['lightPurple'], end_color=colors['lightPurple'], fill_type='solid')
    main['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    main.row_dimensions[2].height = 60

    # Row 3: Spacing
    # Row 4: Instructions
    main.merge_cells('A4:J4')
    main['A4'] = "INSTRUCTIONS: Rate each agent 1-5 (1=Low Priority, 5=Critical)\nWe'll build your top 5 first"
    main['A4'].font = Font(size=12, italic=True)
    main['A4'].fill = PatternFill(start_color=colors['lightBlue'], end_color=colors['lightBlue'], fill_type='solid')
    main['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 5: Spacing

    # Row 6: Column Headers
    headers = ['Priority', 'Area', 'Agent Name', 'What It Does', 'Time Saved/Week',
               'Business Impact', 'Build Complexity', 'Status', 'Your Notes', 'Quick Win?']

    for col_num, header in enumerate(headers, 1):
        cell = main.cell(row=6, column=col_num, value=header)
        cell.font = Font(size=11, bold=True, color=colors['white'])
        cell.fill = PatternFill(start_color=colors['darkGray'], end_color=colors['darkGray'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    widths = [10, 20, 25, 45, 15, 15, 15, 13, 32, 10]
    for col_num, width in enumerate(widths, 1):
        main.column_dimensions[get_column_letter(col_num)].width = width

    # === AGENT DATA ===
    agents_data = []

    # Helper function to add category and agents
    def add_category(category_name, category_color, agents):
        agents_data.append({'type': 'category', 'name': category_name, 'color': category_color})
        for agent in agents:
            agents_data.append({'type': 'agent', 'data': agent})
        agents_data.append({'type': 'spacing'})

    # STRATEGIC INTELLIGENCE
    add_category('🎯 STRATEGIC INTELLIGENCE', colors['purple'], [
        ['', 'Strategic Intelligence', 'Competitive Intelligence Agent', 'Monitors 20+ competitors daily - tracks publications, patents, clinical trials. Alerts on significant developments.', '3 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Strategic Intelligence', 'Research Trend Scanner', 'Identifies emerging cancer research trends 6-12 months before mainstream through citation velocity analysis.', '2 hours', 'HIGH', 'High', 'Not Started', '', True],
        ['', 'Strategic Intelligence', 'IP Landscape Monitor', 'Tracks patent landscape, identifies freedom-to-operate risks and white space opportunities.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Strategic Intelligence', 'Grant Intelligence Agent', 'Finds relevant grants (NIH, NSF, DOD), analyzes winning proposals, estimates success probability.', '3 hours', 'HIGH', 'Medium', 'Not Started', '', True],
    ])

    # INVESTOR RELATIONS
    add_category('💼 INVESTOR RELATIONS', colors['blue'], [
        ['', 'Investor Relations', 'Investor Update Generator', 'Auto-generates weekly/monthly investor updates from research progress, milestones, and achievements.', '4 hours', 'HIGH', 'Low', 'Not Started', '', True],
        ['', 'Investor Relations', 'Pitch Deck Intelligence', 'Keeps pitch deck current with latest milestones, competitive landscape, publications, team accomplishments.', '2 hours', 'MEDIUM', 'Low', 'Not Started', '', True],
        ['', 'Investor Relations', 'Fundraising Opportunity Scanner', 'Identifies potential investors, tracks VC fund raises, suggests timing and warm intro paths.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Investor Relations', 'Grant Writing Assistant', 'Helps write and improve grant proposals based on winning examples and reviewer feedback patterns.', '3 hours', 'HIGH', 'High', 'Not Started', '', False],
    ])

    # RESEARCH OVERSIGHT
    add_category('🔬 RESEARCH OVERSIGHT', colors['green'], [
        ['', 'Research Oversight', 'Breakthrough Detector', 'Flags significant research findings from team before formal reporting. Suggests patent opportunities.', '1 hour', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Research Oversight', 'Research Portfolio Dashboard', 'Real-time view of all projects: status, blockers, dependencies, timeline, risk flags.', '2 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Research Oversight', 'Publication Opportunity Finder', 'Matches research to journals, estimates acceptance likelihood, tracks submission deadlines.', '1 hour', 'MEDIUM', 'Low', 'Not Started', '', False],
        ['', 'Research Oversight', 'Collaboration Matchmaker', 'Identifies external collaboration opportunities, finds complementary research partners.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Research Oversight', 'Research ROI Tracker', 'Tracks cost per publication, grant success rates, program efficiency across all research areas.', '1 hour', 'MEDIUM', 'Low', 'Not Started', '', False],
    ])

    # TEAM MANAGEMENT
    add_category('👥 TEAM MANAGEMENT', colors['orange'], [
        ['', 'Team Management', 'Team Health Monitor', 'Analyzes communication patterns to detect burnout, disengagement before they escalate.', '1 hour', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Team Management', 'Talent Pipeline Agent', 'Monitors top researchers in your field for hiring. Tracks publication records, identifies unhappy researchers.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Team Management', 'Productivity Insights', 'Shows team blockers without micromanaging. Identifies bottlenecks and suggests process improvements.', '1 hour', 'MEDIUM', 'Low', 'Not Started', '', False],
        ['', 'Team Management', 'Onboarding Accelerator', 'Creates personalized onboarding plans for new hires based on role and background.', '1 hour', 'LOW', 'Low', 'Not Started', '', False],
    ])

    # BUSINESS DEVELOPMENT
    add_category('🤝 BUSINESS DEVELOPMENT', colors['cyan'], [
        ['', 'Business Development', 'Partnership Opportunity Scanner', 'Finds pharma/biotech working on complementary research. Identifies partnership fit and warm intros.', '3 hours', 'HIGH', 'High', 'Not Started', '', True],
        ['', 'Business Development', 'Clinical Trial Intelligence', 'Monitors relevant trials, identifies unmet needs, finds trial sponsors and partnership opportunities.', '2 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Business Development', 'Licensing Opportunity Agent', 'Finds in-licensing and out-licensing opportunities. Tracks patent auctions and technology transfers.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Business Development', 'Conference ROI Analyzer', 'Recommends which conferences to attend/sponsor based on attendee analysis and partnership ROI.', '1 hour', 'LOW', 'Low', 'Not Started', '', False],
        ['', 'Business Development', 'Market Intelligence', 'Tracks cancer drug market trends, competitor pipelines, M&A activity, and exit opportunities.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
    ])

    # COMMUNICATIONS & ADMIN
    add_category('📧 COMMUNICATIONS & ADMIN', colors['pink'], [
        ['', 'Communications', 'Email Prioritizer', 'Sorts 200+ daily emails into: urgent/review/delegate/ignore with smart summaries.', '5 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Communications', 'Meeting Prep Agent', 'Prepares briefing docs for every meeting: attendee background, talking points, suggested outcomes.', '3 hours', 'HIGH', 'Low', 'Not Started', '', True],
        ['', 'Communications', 'Board Report Generator', 'Compiles monthly board reports from research progress, financials, team updates automatically.', '4 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Communications', 'Internal Announcements Writer', 'Drafts team communications: milestone celebrations, new hires, policy updates.', '1 hour', 'LOW', 'Low', 'Not Started', '', False],
        ['', 'Communications', 'LinkedIn Content Generator', 'Creates LinkedIn posts highlighting research achievements, team milestones, thought leadership.', '2 hours', 'MEDIUM', 'Low', 'Not Started', '', False],
        ['', 'Communications', 'Press Release Writer', 'Drafts press releases for significant research breakthroughs and company milestones.', '2 hours', 'MEDIUM', 'Low', 'Not Started', '', False],
    ])

    # FINANCIAL OPERATIONS
    add_category('💰 FINANCIAL OPERATIONS', colors['yellow'], [
        ['', 'Financial', 'Budget Optimizer', 'Recommends resource reallocation based on research progress and ROI analysis.', '2 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Financial', 'Burn Rate Monitor', 'Tracks spending velocity daily, alerts on budget risks, calculates runway.', '1 hour', 'HIGH', 'Low', 'Not Started', '', True],
        ['', 'Financial', 'Research ROI Analyzer', 'Calculates cost per publication, grant success ROI, program efficiency. Investment recommendations.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Financial', 'Vendor Intelligence', 'Monitors equipment/service vendors for better pricing, tracks contract renewals, suggests alternatives.', '1 hour', 'LOW', 'Low', 'Not Started', '', False],
    ])

    # REGULATORY & COMPLIANCE
    add_category('⚖️ REGULATORY & COMPLIANCE', colors['red'], [
        ['', 'Regulatory', 'Regulatory Intelligence', 'Monitors FDA/regulatory changes affecting ACM research. Tracks approval trends and competitor approvals.', '2 hours', 'MEDIUM', 'Medium', 'Not Started', '', False],
        ['', 'Regulatory', 'Risk Monitor', 'Flags compliance risks, research ethics issues, safety concerns, IP infringement risks.', '1 hour', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Regulatory', 'Audit Preparation', 'Maintains audit-ready documentation, flags potential audit issues before they become problems.', '1 hour', 'MEDIUM', 'Low', 'Not Started', '', False],
    ])

    # PERSONAL PRODUCTIVITY
    add_category('🧠 PERSONAL PRODUCTIVITY', colors['indigo'], [
        ['', 'Personal', 'Decision Intelligence', 'Summarizes complex issues with pros/cons, risk assessment, data-driven recommendations.', '3 hours', 'HIGH', 'High', 'Not Started', '', True],
        ['', 'Personal', 'Reading Digest Agent', 'Curates must-read papers, industry news, competitor updates into 10-minute daily digest.', '5 hours', 'HIGH', 'Medium', 'Not Started', '', True],
        ['', 'Personal', 'Calendar Optimizer', 'Suggests meeting consolidation, blocks focus time, identifies unnecessary meetings.', '2 hours', 'MEDIUM', 'Low', 'Not Started', '', True],
        ['', 'Personal', 'Travel Coordinator', 'Books travel, manages itineraries, prepares trip briefs with meeting schedules and local intel.', '2 hours', 'LOW', 'Low', 'Not Started', '', False],
    ])

    # Write data to sheet
    row = 7
    first_data_row = 7

    for item in agents_data:
        if item['type'] == 'category':
            main.merge_cells(f'A{row}:J{row}')
            cell = main[f'A{row}']
            cell.value = item['name']
            cell.font = Font(size=12, bold=True, color=colors['white'])
            cell.fill = PatternFill(start_color=item['color'], end_color=item['color'], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        elif item['type'] == 'agent':
            for col_num, value in enumerate(item['data'], 1):
                cell = main.cell(row=row, column=col_num, value=value)
                # Wrap text for description and notes
                if col_num in [4, 9]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                # Center alignment for specific columns
                if col_num in [1, 5, 6, 7, 8, 10]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        elif item['type'] == 'spacing':
            row += 1

    last_data_row = row - 1

    # Add data validation
    # Priority dropdown
    priority_dv = DataValidation(type="list", formula1='"5 - Critical,4 - High,3 - Medium,2 - Low,1 - Not Now"', allow_blank=True)
    main.add_data_validation(priority_dv)
    priority_dv.add(f'A{first_data_row}:A{last_data_row}')

    # Business Impact dropdown
    impact_dv = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW"', allow_blank=True)
    main.add_data_validation(impact_dv)
    impact_dv.add(f'F{first_data_row}:F{last_data_row}')

    # Build Complexity dropdown
    complexity_dv = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
    main.add_data_validation(complexity_dv)
    complexity_dv.add(f'G{first_data_row}:G{last_data_row}')

    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Not Started,Planning,In Progress,Complete"', allow_blank=True)
    main.add_data_validation(status_dv)
    status_dv.add(f'H{first_data_row}:H{last_data_row}')

    # === SUMMARY DASHBOARD (L1:N20) ===
    main.merge_cells('L1:N1')
    dash_header = main['L1']
    dash_header.value = '📊 SUMMARY DASHBOARD'
    dash_header.font = Font(size=14, bold=True, color=colors['white'])
    dash_header.fill = PatternFill(start_color=colors['darkGray'], end_color=colors['darkGray'], fill_type='solid')
    dash_header.alignment = Alignment(horizontal='center', vertical='center')

    main['L3'] = 'Total Agents:'
    main['M3'] = f'=COUNTA(C{first_data_row}:C{last_data_row})'

    main['L5'] = 'Rated by You:'
    main['M5'] = f'=COUNTA(A{first_data_row}:A{last_data_row})'

    main['L6'] = 'Avg Priority:'
    main['M6'] = f'=AVERAGE(A{first_data_row}:A{last_data_row})'

    main['L8'] = 'TIME SAVINGS:'
    main['L8'].font = Font(bold=True)
    main['L9'] = 'Quick Wins:'
    main['M9'] = '25 hrs/week'
    main['L10'] = 'Total Possible:'
    main['M10'] = '67 hrs/week'

    main['L12'] = 'COMPLEXITY:'
    main['L12'].font = Font(bold=True)
    main['L13'] = 'Low:'
    main['M13'] = f'=COUNTIF(G{first_data_row}:G{last_data_row},"Low")'
    main['L14'] = 'Medium:'
    main['M14'] = f'=COUNTIF(G{first_data_row}:G{last_data_row},"Medium")'
    main['L15'] = 'High:'
    main['M15'] = f'=COUNTIF(G{first_data_row}:G{last_data_row},"High")'

    main['L17'] = 'TOP 5 PRIORITIES:'
    main['L17'].font = Font(bold=True)
    main['L18'] = '1.'
    main['M18'] = f'=IFERROR(INDEX(C{first_data_row}:C{last_data_row}, MATCH(MAX(A{first_data_row}:A{last_data_row}), A{first_data_row}:A{last_data_row}, 0)), "")'
    main['L19'] = '2.'
    main['M19'] = f'=IFERROR(INDEX(C{first_data_row}:C{last_data_row}, MATCH(LARGE(A{first_data_row}:A{last_data_row},2), A{first_data_row}:A{last_data_row}, 0)), "")'
    main['L20'] = '3.'
    main['M20'] = f'=IFERROR(INDEX(C{first_data_row}:C{last_data_row}, MATCH(LARGE(A{first_data_row}:A{last_data_row},3), A{first_data_row}:A{last_data_row}, 0)), "")'

    # Set dashboard column widths
    main.column_dimensions['L'].width = 18
    main.column_dimensions['M'].width = 25
    main.column_dimensions['N'].width = 15

    # === QUICK WINS SHEET ===
    quick.merge_cells('A1:J1')
    quick['A1'] = '⚡ QUICK WIN AGENTS\nHigh Impact + Fast to Build'
    quick['A1'].font = Font(size=24, bold=True, color=colors['white'])
    quick['A1'].fill = PatternFill(start_color=colors['amber'], end_color=colors['amber'], fill_type='solid')
    quick['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    quick.row_dimensions[1].height = 72

    # Headers
    for col_num, header in enumerate(headers, 1):
        cell = quick.cell(row=3, column=col_num, value=header)
        cell.font = Font(size=11, bold=True, color=colors['white'])
        cell.fill = PatternFill(start_color=colors['darkGray'], end_color=colors['darkGray'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    for col_num, width in enumerate(widths, 1):
        quick.column_dimensions[get_column_letter(col_num)].width = width

    # Add quick win agents (filter by Quick Win? = True)
    quick_row = 4
    for item in agents_data:
        if item['type'] == 'agent' and item['data'][9] == True:
            for col_num, value in enumerate(item['data'], 1):
                cell = quick.cell(row=quick_row, column=col_num, value=value)
                if col_num in [4, 9]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col_num in [1, 5, 6, 7, 8, 10]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            quick_row += 1

    # === BUILD ROADMAP SHEET ===
    roadmap.merge_cells('A1:I1')
    roadmap['A1'] = '🚀 AGENT BUILD ROADMAP'
    roadmap['A1'].font = Font(size=24, bold=True, color=colors['white'])
    roadmap['A1'].fill = PatternFill(start_color=colors['purpleGrad'], end_color=colors['purpleGrad'], fill_type='solid')
    roadmap['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    r_headers = ['Phase', 'Agent Name', 'Priority Score', 'Time Saved', 'Build Time', 'Dependencies', 'Start Date', 'Launch Date', 'Owner']
    for col_num, header in enumerate(r_headers, 1):
        cell = roadmap.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True, color=colors['white'])
        cell.fill = PatternFill(start_color=colors['darkGray'], end_color=colors['darkGray'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Phase sections
    roadmap.merge_cells('A4:I4')
    roadmap['A4'] = 'PHASE 1: FOUNDATIONS (Weeks 1-4)\nQuick wins with immediate impact'
    roadmap['A4'].font = Font(bold=True, color=colors['white'])
    roadmap['A4'].fill = PatternFill(start_color=colors['blue'], end_color=colors['blue'], fill_type='solid')
    roadmap['A4'].alignment = Alignment(wrap_text=True)

    roadmap.merge_cells('A10:I10')
    roadmap['A10'] = 'PHASE 2: INTELLIGENCE (Weeks 5-12)\nStrategic and analytical agents'
    roadmap['A10'].font = Font(bold=True, color=colors['white'])
    roadmap['A10'].fill = PatternFill(start_color=colors['cyan'], end_color=colors['cyan'], fill_type='solid')
    roadmap['A10'].alignment = Alignment(wrap_text=True)

    roadmap.merge_cells('A16:I16')
    roadmap['A16'] = 'PHASE 3: AUTOMATION (Weeks 13-24)\nProcess optimization agents'
    roadmap['A16'].font = Font(bold=True, color=colors['white'])
    roadmap['A16'].fill = PatternFill(start_color=colors['green'], end_color=colors['green'], fill_type='solid')
    roadmap['A16'].alignment = Alignment(wrap_text=True)

    # Column widths for roadmap
    r_widths = [15, 38, 18, 15, 15, 28, 15, 15, 20]
    for col_num, width in enumerate(r_widths, 1):
        roadmap.column_dimensions[get_column_letter(col_num)].width = width

    # Save workbook
    output_path = os.path.join(output_dir, 'Agent_Portfolio.xlsx')
    wb.save(output_path)
    print(f"✅ Excel file created successfully: {output_path}")
    return output_path

if __name__ == '__main__':
    output_path = create_agent_portfolio_excel()
    print(f"\n📊 Agent Portfolio Excel file has been created!")
    print(f"📁 Location: {os.path.abspath(output_path)}")
